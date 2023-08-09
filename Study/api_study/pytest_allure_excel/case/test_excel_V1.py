# coding=utf-8
import os
import allure
import openpyxl
import pytest
from Study.api_study.api_keyword.test_apikey import Apikey
from Study.api_study.pytest_allure_excel.data_driver.excel_driver import load_excel

# 在当前文件中所有用例执行之前运行一次
def setup_module():
    # 1.定义全局变量
    global  ak,wb,sheet,path
    # 2.实例化工具
    ak = Apikey()
    # 3.初始化excel文件
    path = '../test_data/api_cases_V1.xlsx'
    wb = openpyxl.load_workbook(path)
    # 注意页签名称开头大小写
    sheet = wb['Sheet1']

@allure.title('登录')
@pytest.mark.parametrize('userData',load_excel())
def test_01(userData):
    # 行数
    r = userData[0] + 1
    ak = Apikey()
    with allure.step('01.登录'):
        try:
            dict_data = {
                "url" : userData[1] + userData[2],
                "params" : eval(userData[4]),
                "headers" : eval(userData[5]),
                userData[7] : eval(userData[6])
            }
            print(dict_data)
        except:
            print('====================实际结果====================')
            print('请求参数有误，请检查')
            sheet.cell(r,11).value = '请求参数有误，请检查'
            wb.save(path)
        # 反射
        res = getattr(ak,userData[3])(**dict_data)
        print(res.text)
        # 结果校验
        try:
            # 实际结果
            result = None
            result = ak.get_text(res.text,userData[8])
            print(result == userData[9])
            if result == userData[9]:
                sheet.cell(r,11).value = '通过'
            else:
                sheet.cell(r,11).value = '失败'
            wb.save(path)
        except:
            print('====================实际结果====================')
            print('预期结果的jsonpath表达式有误,请检查')
            sheet.cell(r, 11).value = '预期结果的jsonpath表达式有误,请检查'
            wb.save(path)
        finally:
            assert result == userData[9]
if __name__ == '__main__':
    pytest.main(['-v', 'test_excel_V1.py', '--alluredir', '../result', '--clean-alluredir'])
    os.system('allure generate ../result -o ../report --clean')

