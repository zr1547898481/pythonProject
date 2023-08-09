# coding=utf-8
import os
import allure
import openpyxl
import pytest
from Study.api_study.api_keyword.test_apikey import Apikey
from Study.api_study.pytest_allure_excel.VAR.var import *
from Study.api_study.pytest_allure_excel.data_driver.excel_driver import load_excel


# 在当前文件中所有用例执行之前运行一次
def setup_module():
    # 1.定义全局变量
    global ak,wb,sheet,path,all_var
    # 2.实例化工具类
    ak = Apikey()
    # 3.初始化excel文件
    path = EXCEL_PATH4
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    # 4.定义字典储存jsonpath表达式
    all_var = {}

@allure.title('登录')
@pytest.mark.parametrize('data',load_excel())
def test_01(data):
    # 动态加载用例标题
    if data[11]  is not None:
        allure.dynamic.title(data[11])
    # 用例功能点
    if data[16]  is not None:
        allure.dynamic.story(data[16])
    # 用例模块
    if data[17]  is not None:
        allure.dynamic.feature(data[17])
    # 用例等级
    if data[18]  is not None:
        allure.dynamic.severity(data[18])
    # 用例描述
    if data[19]  is not None:
        allure.dynamic.description(data[19])
    # 行数
    r = data[0] + 1

    # json数据解析
    try:
        dict_data = {
            "url" : data[1] + data[2],
            "params" : eval(data[4]),
            "headers" : eval(data[5]),
            data[7] : eval(data[6])
        }
    except:
        print('请求参数错误,请检查')
        # 将结果写入excel文件
        sheet.cell(r,11).value = '请求参数错误,请检查'
        # 保存文件
        wb.save(path)

    # 反射发起请求
    res = getattr(ak,data[3])(**dict_data)
    print(res.text)

    # 获取json提取值变量名称
    if data[12]  is not None:
        var_name = data[12]
        name_list = var_name.split(';')
    # 获取jsonpath表达式
    if data[13] is not None:
        var_json = data[13]
        json_list =var_json.split(';')

    for i in range(len(name_list)) :
        # 通过循环遍历获取遍历名
        key = name_list[i]
        # 通过遍历获取jsonpath表达式
        jsonExp = json_list[i]
        # 动态传入jsonpath表达式，获取到对应的value
        json_value = ak.get_text(res.text,jsonExp)
        # key为json提取值的引用变量名，value为提取的值
        all_var[key] = json_value
    print(all_var)

    # 接口返回结果校验
    try :
        print('===============实际结果===============')
        result = None
        result = ak.get_text(res.text,data[8])
        print(result == data[9])
        if result == data[9] :
            # 列数从0开始
            sheet.cell(r,11,value = '通过')
        else:
            sheet.cell(r,11,value = '失败')
        wb.save(path)

    except:
        print('jsonpath表达式有误,请检查')
        sheet.cell(r,11).value = 'jsonpath表达式有误,请检查'
        wb.save(path)

    finally:
        assert result ==data[9]

    # 存在数据库结果校验,并且需要参数关联
    if data[21] is not None:
        try:
            # 获取sql
            str_sql = data[20]
            sqlVar = data[21]
            # 分割数据库变量
            sqlVar_list = sqlVar.split(';')
            # 存放关联参数变量名值
            list1 = []
            for i in range(len(sqlVar_list)):
                # 通过下标遍历数据库变量值，去除外层引号添加到空list中
                list1.append(eval(sqlVar_list[i]))
            # 完整sql
            sql = str_sql.format(*list1)
            # sql查询结果
            sql_result = None
            sql_result = ak.check_sqldb(sql)
            if sql_result == data[22]:
                sheet.cell(r,11).value = '通过'
            else:
                sheet.cell(r,11).value = '不通过'
            wb.save(path)

        except Exception as e:
            print(f'报错信息为:{e}')
            print('sql有误,请检查')
            sheet.cell(r, 11).value = 'sql有误,请检查'
            wb.save(path)

        finally:
            assert sql_result == data[22]
    # 存在数据库结果校验,不需要参数关联
    elif data[20] is not None:
        try:
            # sql查询结果
            sql = data[20]
            sql_result = None
            sql_result = ak.check_sqldb(sql)
            if sql_result == data[22]:
                sheet.cell(r, 11).value = '通过'
            else:
                sheet.cell(r, 11).value = '不通过'
            wb.save(path)

        except:
            print('sql有误,请检查')
            sheet.cell(r, 11).value = 'sql有误,请检查'
            wb.save(path)

        finally:
            assert sql_result == data[22]


if __name__ == '__main__':
    pytest.main(['-v','test_excel_V4.py','--alluredir','../result','--clean-alluredir'])
    os.system('allure generate ../result -o ../report --clean')
